import openpyxl as xl
##import chart
from openpyxl.chart import BarChart, Reference

def process_workbook(filename):

 
    wb = (xl.load_workbook(filename))

    ##access sheet 1 in your workbook (sheet having the data)
    sheet = wb['Sheet1']

    ##access column/row of cell (ex: transaction_id)
    # cell = sheet.cell(1,1) 

    #iterate each row and get values in third column (price)
    #$tart row from 2 b/c dont want to include heading titles in our calculation
    #multiply each value by 0.9 and add them to new column
    for row in range(2, sheet.max_row+1):
        cell = sheet.cell(row, 3)
        corrected_price = cell.value * 0.9
        ##access fourth column to put corrected values
        corrected_price_cell = sheet.cell(row,4)
        corrected_price_cell.value = corrected_price

    ##select values in 4th column ONLY and create chart
    values = Reference(sheet, 
            min_row=2, 
            max_row=sheet.max_row,
            min_col=4,
            max_col=4)
    chart = BarChart()
    chart.add_data(values)

    ##create and decided where to add  ('e2' cell)
    sheet.add_chart(chart, 'e2')

    ##you should see your chart in excel sheet
    wb.save(filename)

